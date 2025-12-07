from flask import Blueprint, url_for, redirect, request, jsonify, render_template, session
from datetime import datetime, timedelta
import base64
from openpyxl import load_workbook
from helpers import *
from db import *
import base64
from inr import inr

page = Blueprint('page', __name__)


saree_materials = [
    "Banarasi Silk", "Kanjivaram Silk", "Tussar Silk", "Mysore Silk",
    "Handloom Cotton", "Chanderi Cotton", "Tant Cotton",
    "Georgette", "Chiffon", "Crepe", "Net", 
    "Linen", "Satin", "Organza", "Velvet"
]

creds = {'Tulsiyan@rootUser': 'password'}


@page.route("/")
def index():
    if session.get('user') == None:
        return redirect('/login')

    user_count = fetch_ttl_users()
    sku_count = fetch_sku_count()

    return render_template('index.html', page_name="homepage", usr_count=user_count, sku_count=sku_count)


@page.route("/add", methods = ['POST', 'GET'])
def add():
    if session.get('user') == None:
        return redirect('/login')

    if request.method == "GET":
        return render_template('add.html', page_name="add products", material=saree_materials)
    else:
        data = request.form

        skuid = product_handler.create_sku()
        title = data.get('title').strip()
        vendorid = data.get('vid', 'NULL').strip()
        product_desc = data.get('desc').strip()
        product_kwords = data.get('keywords').split(',')
        product_weight = data.get('weight').strip().strip()
        original_price = data.get('price').replace(',', '').strip()
        discounted_price = data.get('dsc-price').replace(',', '').strip()
        product_stock = data.get('stock').strip().strip()
        slen = data.get('slen').strip().strip()
        blen = data.get('blen').strip().strip()
        material = data.get('material').strip().strip()
        color = data.get('color').strip()
        care = data.get('product-care').strip().strip()
        art = data.get('art').strip()
        stitch = data.get('stitch').strip()

        file01 = request.files.get('img01')
        file02 = request.files.get('img02')
        file03 = request.files.get('img03')
        file04 = request.files.get('img04')

        main_image = file01.read()
        img02 = file02.read()
        img03 = file03.read()
        img04 = file04.read()

        date = datetime.now()
        date = date.strftime("%Y-%m-%d")
        add_product(skuid, vendorid, title, product_kwords, original_price, discounted_price, product_weight, product_stock, product_desc,
                    slen, blen, material, color, care, art, stitch, date, main_image, img02, img03, img04)
        return redirect('/add')
    

@page.route("/draft", methods=['POST', 'GET'])
def draft():
    if session.get('user') == None:
        return redirect('/login')
    
    if request.method == "GET":
        return render_template('draft.html', page_name="draft products")
    


@page.route("/edit", methods=['POST', 'GET'])
def edit():
    if session.get('user') == None:
        return redirect('/login')
    
    if request.method == "GET":
        return render_template('edit.html', page_name="edit products", material=saree_materials)
    else:
        skuid = session.get('skuid')
        if request.content_type == 'application/json':
            query = request.get_json().get('query')
            session['skuid'] = query
            
            # write a function to fetch the provided query and fetch the product data and return
            cursor = mysql.connection.cursor()
            data = cursor.execute('''
                            select vendor_id, product_desc, original_price, disc_price, product_weight_gm, 
                       product_stock, product_details, saree_len, blouse_len, product_material, product_color, product_care, product_image, 
                       product_img01, product_img02, product_img03, product_art, product_stitch
                       from inventory where skuID=%s
                       ''', (query, ))

            data = cursor.fetchone()
            cursor.close()
            # print(data)
            product_details = {
                'img01': base64.b64encode(data[12]).decode('utf-8') if data[11] else None,
                'img02': base64.b64encode(data[13]).decode('utf-8') if data[12] else None,
                'img03': base64.b64encode(data[14]).decode('utf-8') if data[13] else None,
                'img04': base64.b64encode(data[15]).decode('utf-8') if data[14] else None,
                'title': data[1],
                'vid': data[0],
                'desc': data[6],
                'weight': data[4],
                'og_price': data[2],
                'disc_price': data[3],
                'stock': data[5],
                'slen': data[7],
                'blen': data[8],
                'material': data[9],
                'color': data[10],
                'care': data[11],
                'art': data[16],
                'stitch': data[17]
            }

            cursor = mysql.connection.cursor()
            cursor.execute('''
                           select search_result from searching_keywords where skuId = %s
                           ''', (query, ))

            keywords = cursor.fetchall()
            cursor.close()

            keywords = [ item[0]
                for item in keywords
            ]
            # print(keywords)
            response_data = {'data': product_details, 'keywords': (', ').join(keywords)}
            # print(response_data)
            return jsonify(response_data)

        # when the update buttons is clicked update the data
        data = request.form
        title = data.get('title').strip()
        vendorid = data.get('vid', 'NULL').strip()
        product_desc = data.get('desc').strip()
        product_kwords = data.get('keywords').split(',')
        product_weight = data.get('weight').strip().strip()
        original_price = data.get('price').replace(',', '').strip()
        discounted_price = data.get('dsc-price').replace(',', '').strip()
        product_stock = data.get('stock').strip().strip()
        slen = data.get('slen').strip().strip()
        blen = data.get('blen').strip().strip()
        material = data.get('material').strip().strip()
        color = data.get('color').strip()
        care = data.get('product-care').strip().strip()
        art = data.get('art').strip()
        stitch= data.get('stitch').strip()

        file01 = request.files['img01']
        main_image = file01.read()

        file02 = request.files['img02']
        img02 = file02.read()
        
        file03 = request.files['img03']
        img03 = file03.read()
        
        file04 = request.files['img04']
        img04 = file04.read()

        date = datetime.now()
        date = date.strftime("%Y-%m-%d")
        cursor = mysql.connection.cursor()
        cursor.execute('''
                       update inventory set
                       product_desc = %s,
                       vendor_id = %s,
                       product_details = %s, 
                       product_weight_gm = %s,
                       original_price = %s, 
                       disc_price = %s, 
                       saree_len = %s, 
                       blouse_len = %s,
                       product_material = %s,
                       product_color = %s,
                       product_care = %s,
                       product_art = %s,
                       product_stitch = %s,
                       updation_date = %s
                       where skuID = %s
        ''', (title, vendorid, product_desc, product_weight, original_price, discounted_price, slen, blen, material, color, care, art, stitch, date, skuid))
        
        if file01:
            comp_img = compress_image(main_image)
            com_main_img = compress_main_image(main_image)
            cursor.execute('''update images
                            set img1 = %s
                            where skuID = %s
                            ''', (com_main_img, skuid))
            
            cursor.execute('''update inventory
                            set product_image = %s
                            where skuID = %s
                            ''', (comp_img, skuid))

        if file02:
            comp_img = compress_image(img02)
            com_main_img = compress_main_image(img02)
            cursor.execute('''update images
                            set 
                           img2 = %s
                            where skuID = %s
                            ''', (com_main_img, skuid))
            
            cursor.execute('''update inventory
                            set
                            product_img01 = %s
                            where skuID = %s
                            ''', (comp_img, skuid))
            
        if file03:
            comp_img = compress_image(img03)
            com_main_img = compress_main_image(img03)
            cursor.execute('''update images
                            set 
                            img3 = %s
                            where skuID = %s
                            ''', (com_main_img, skuid))
            
            cursor.execute('''update inventory
                            set
                            product_img02 = %s
                            where skuID = %s
                            ''', (comp_img, skuid))
            
        if file04:
            comp_img = compress_image(img04)
            com_main_img = compress_main_image(img04)
            cursor.execute('''update images
                            set 
                            img4 = %s
                            where skuID = %s
                            ''', (com_main_img, skuid))
            
            cursor.execute('''update inventory
                            set
                            product_img03 = %s
                            where skuID = %s
                            ''', (comp_img, skuid))

        # clearing all the prev searching keywords
        cursor.execute('''
                            delete from searching_keywords where skuID = %s''', (skuid, ))

        # fetching the already existing productids
        cursor.execute('''
                        select count(productID) from products where skuID = %s
                       ''', (skuid, ))
        
        existing_productids = cursor.fetchone()[0]
        
        for _ in product_kwords:
            keyword = _.strip()

            cursor.execute('''insert into searching_keywords(skuID, search_result) values(%s, %s)''',
                           (skuid, keyword))
        
        if int(product_stock) > existing_productids:
            cursor.execute('''update inventory
                           set product_stock=%s
                           where skuID=%s''', (product_stock, skuid))

            for _ in range(int(product_stock) - existing_productids):
                productid = product_handler.create_productid()
                cursor.execute('''insert into products(skuID, productID) values(%s, %s)''', (skuid, productid))

        mysql.connection.commit()
        cursor.close()
        print('product updated successfully *_*')
        return redirect('/edit')
    

@page.route('/login', methods=['POST', 'GET'])
def login():
    if request.method == 'POST':
        data = request.form
        empl_code = data.get('empl-code')
        psswrd = data.get('password')

        if empl_code and psswrd:
            if creds[empl_code] != psswrd:
                return redirect('/login')
            else:
                session['user'] = empl_code
                return redirect('/')
        else:
            return redirect('/login')
        
    else:
        return render_template('login.html', page_name='login')
    

@page.route('/csv', methods=['POST', 'GET'])
def csv_upload():
    if request.method == 'GET':
        return render_template('csv.html', page_name='csv upload')
    else:
        file = request.files.get('file')

        if file.filename == '':
            return "No selected file", 400

        # extension verification
        filename = str(file.filename)
        ext = filename.split('.')[1]
        if (ext != 'xlsx'):
            return "invalid file or file name has (.) in it"


        wb = load_workbook(filename=file)  #load workbook or xlsx file
        ws = wb.active
        # Access all images
        for sheetname in wb.sheetnames:
            ws = wb[sheetname]

        skus = []   # list of skus to return and show back on the screen
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i > 0:
                skuid = product_handler.create_sku()
                skus.append(skuid)
                title = row[0]
                vendorid = row[1]
                product_desc = row[2]
                product_kwords = row[3].split(', ')
                product_weight = row[4]
                original_price = row[5]
                discounted_price = row[6]
                product_stock = row[7]
                slen = row[8]
                blen = row[9]
                material = row[10], 
                color = row[11]
                care = row[12]
                art = row[13]
                stitch = row[14]
                date = datetime.now().date()
                print(skuid, vendorid, title, product_kwords, original_price, discounted_price, product_weight, product_stock, product_desc,
                    slen, blen, material, color, care, art, stitch, date)
                add_product(skuid, vendorid, title, product_kwords, original_price, discounted_price, product_weight, product_stock, product_desc,
                    slen, blen, material, color, care, art, stitch, date)
                # print(row)

        return skus
    

@page.route('/delivery', methods=['POST', 'GET'])
def delivery():
    if request.method == 'GET':
        cursor = mysql.connection.cursor()
        cursor.execute('''
                        select orderID, skuID, productID 
                        from ready_for_delivery
                        ''')
        data = cursor.fetchall()
        cursor.close()
        print(data)
        return render_template('delivery.html', page_name='process delivery', data=data)
    return

@page.route('/customers')
def fetch_all_cx():
    cursor = mysql.connection.cursor()
    cursor.execute('''SELECT userID, user_name,user_number FROM user''')
    data = cursor.fetchall()
    cursor.close()
    print("Fetched users: ", data)
    return render_template('customers.html',data=data,page_name='customers')

@page.route('/sku')
def fetch_all_sku():
    cursor = mysql.connection.cursor()
    cursor.execute("SELECT skuID, product_desc, disc_price, product_image FROM inventory")
    data = cursor.fetchall()
    cursor.close()

    data= [{
        'skuid': _[0],
        'title': _[1],
        'selling_price': inr(_[2]).formate(),
        'image': base64.b64encode(_[3]).decode('utf-8')
        }   for _ in data
    ]
    # print("Fetched SKUs:", data)
    return render_template('sku.html', data=data, page_name='sku')

